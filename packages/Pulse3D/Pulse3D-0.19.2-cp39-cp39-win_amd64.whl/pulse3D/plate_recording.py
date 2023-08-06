# -*- coding: utf-8 -*-
import datetime
import glob
import os
import tempfile
from typing import Any
from typing import Optional
import uuid
import zipfile

import h5py
from nptyping import NDArray
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from semver import VersionInfo
from xlsxwriter.utility import xl_cell_to_rowcol

from .compression_cy import compress_filtered_magnetic_data
from .constants import *
from .transforms import apply_empty_plate_calibration
from .transforms import apply_noise_filtering
from .transforms import apply_sensitivity_calibration
from .transforms import calculate_displacement_from_voltage
from .transforms import calculate_force_from_displacement
from .transforms import calculate_voltage_from_gmr
from .transforms import create_filter
from .transforms import noise_cancellation


def _get_col_as_array(sheet: Worksheet, zero_based_row: int, zero_based_col: int) -> NDArray[(2, Any), float]:
    col_array = []
    result = _get_cell_value(sheet, zero_based_row, zero_based_col)
    zero_based_row += 1
    while result:
        col_array.append(float(result))
        result = _get_cell_value(sheet, zero_based_row, zero_based_col)
        zero_based_row += 1
    return np.array(col_array)


def _get_single_sheet(file_name: str) -> Any:
    work_book = load_workbook(file_name)
    return work_book[work_book.sheetnames[0]]


def _get_cell_value(sheet: Worksheet, zero_based_row: int, zero_based_col: int) -> Optional[str]:
    result = sheet.cell(row=zero_based_row + 1, column=zero_based_col + 1).value
    if result is None:
        return result
    return str(result)


def _get_excel_metadata_value(sheet: Worksheet, metadata_uuid: uuid.UUID) -> Optional[str]:
    """Return a user-entered metadata value."""
    metadata_description = METADATA_UUID_DESCRIPTIONS[metadata_uuid]
    cell_name = EXCEL_OPTICAL_METADATA_CELLS.get(metadata_uuid, None)
    if cell_name is None:
        raise NotImplementedError(
            f"Metadata value for {metadata_description} is not contained in excel files of well data"
        )
    row, col = xl_cell_to_rowcol(cell_name)
    result = _get_cell_value(sheet, row, col)
    if result is None and metadata_uuid != INTERPOLATION_VALUE_UUID:
        raise MetadataNotFoundError(f"Metadata entry not found for {metadata_description}")
    return result


def _load_optical_file_attrs(sheet: Worksheet):
    raw_tissue_reading = np.array((_get_col_as_array(sheet, 1, 0), _get_col_as_array(sheet, 1, 1)))

    value = _get_excel_metadata_value(sheet, TISSUE_SAMPLING_PERIOD_UUID)
    if value is None:
        raise NotImplementedError(
            "Tissue Sampling Period should never be None here. A MetadataNotFoundError should have been raised by get_excel_metadata_value"
        )
    sampling_period = int(round(1 / float(value), 6) * MICRO_TO_BASE_CONVERSION)

    interpolation_value = _get_excel_metadata_value(sheet, INTERPOLATION_VALUE_UUID)
    if interpolation_value is None:
        interpolation_value = sampling_period
    else:
        interpolation_value = float(interpolation_value) * MICRO_TO_BASE_CONVERSION

    begin_recording = _get_excel_metadata_value(sheet, UTC_BEGINNING_RECORDING_UUID)
    begin_recording = datetime.datetime.strptime(begin_recording, "%Y-%m-%d %H:%M:%S")

    twenty_four_well = LabwareDefinition(row_count=4, column_count=6)
    well_name = _get_excel_metadata_value(sheet, WELL_NAME_UUID)

    attrs = {
        FILE_FORMAT_VERSION_METADATA_KEY: "0.1.1",
        TISSUE_SENSOR_READINGS: raw_tissue_reading,
        REFERENCE_SENSOR_READINGS: np.zeros(raw_tissue_reading.shape),
        str(INTERPOLATION_VALUE_UUID): interpolation_value,
        str(TISSUE_SAMPLING_PERIOD_UUID): sampling_period,
        str(UTC_BEGINNING_RECORDING_UUID): begin_recording,
        str(MANTARRAY_SERIAL_NUMBER_UUID): _get_excel_metadata_value(sheet, MANTARRAY_SERIAL_NUMBER_UUID),
        str(PLATE_BARCODE_UUID): _get_excel_metadata_value(sheet, PLATE_BARCODE_UUID),
        str(WELL_NAME_UUID): well_name,
        str(WELL_INDEX_UUID): twenty_four_well.get_well_index_from_well_name(well_name),
    }

    return attrs


class WellFile:
    def __init__(self, file_path: str, sampling_period=None):
        if file_path.endswith(".h5"):
            self.file = h5py.File(file_path, "r")
            self.file_name = os.path.basename(self.file.filename)

            self.is_force_data = True
            self.is_magnetic_data = True

            self.attrs = {attr: self.file.attrs[attr] for attr in list(self.file.attrs)}
            self.version = self[FILE_FORMAT_VERSION_METADATA_KEY]

            # extract datetime
            self[UTC_BEGINNING_RECORDING_UUID] = self._extract_datetime(UTC_BEGINNING_RECORDING_UUID)
            self[UTC_BEGINNING_DATA_ACQUISTION_UUID] = self._extract_datetime(
                UTC_BEGINNING_DATA_ACQUISTION_UUID
            )
            self[UTC_FIRST_TISSUE_DATA_POINT_UUID] = self._extract_datetime(UTC_FIRST_TISSUE_DATA_POINT_UUID)
            self[UTC_FIRST_REF_DATA_POINT_UUID] = self._extract_datetime(UTC_FIRST_REF_DATA_POINT_UUID)

        elif file_path.endswith(".xlsx"):
            self._excel_sheet = _get_single_sheet(file_path)
            self.file_name = os.path.basename(file_path)
            self.attrs = {k: v for (k, v) in _load_optical_file_attrs(self._excel_sheet).items()}
            self.version = self[FILE_FORMAT_VERSION_METADATA_KEY]

            self.is_magnetic_data = False
            self.is_force_data = (
                "y" in str(_get_excel_metadata_value(self._excel_sheet, TWITCHES_POINT_UP_UUID)).lower()
            )

        # setup noise filter
        self.tissue_sampling_period = (
            sampling_period if sampling_period else self[TISSUE_SAMPLING_PERIOD_UUID]
        )
        self.noise_filter_uuid = (
            TSP_TO_DEFAULT_FILTER_UUID[self.tissue_sampling_period] if self.is_magnetic_data else None
        )
        self.filter_coefficients = (
            create_filter(self.noise_filter_uuid, self.tissue_sampling_period)
            if self.noise_filter_uuid
            else None
        )

        is_untrimmed = self.get(IS_FILE_ORIGINAL_UNTRIMMED_UUID, True)
        time_trimmed = None if is_untrimmed else self.attrs[TRIMMED_TIME_FROM_ORIGINAL_START_UUID]

        if self.is_magnetic_data:
            # load sensor data
            self[TISSUE_SENSOR_READINGS] = self._load_reading(TISSUE_SENSOR_READINGS, time_trimmed)
            self[REFERENCE_SENSOR_READINGS] = self._load_reading(REFERENCE_SENSOR_READINGS, time_trimmed)

        self._load_magnetic_data()

    def _load_magnetic_data(self):
        adj_raw_tissue_reading = self[TISSUE_SENSOR_READINGS].copy()
        f = MICROSECONDS_PER_CENTIMILLISECOND if self.is_magnetic_data else MICRO_TO_BASE_CONVERSION
        adj_raw_tissue_reading[0] *= f

        # magnetic data is flipped
        if self.is_magnetic_data:
            adj_raw_tissue_reading[1] *= -1

        self.raw_tissue_magnetic_data: NDArray[(2, Any), int] = adj_raw_tissue_reading
        self.raw_reference_magnetic_data: NDArray[(2, Any), int] = self[REFERENCE_SENSOR_READINGS].copy()

        self.sensitivity_calibrated_tissue_gmr: NDArray[(2, Any), int] = apply_sensitivity_calibration(
            self.raw_tissue_magnetic_data
        )

        self.sensitivity_calibrated_reference_gmr: NDArray[(2, Any), int] = apply_sensitivity_calibration(
            self.raw_reference_magnetic_data
        )

        self.noise_cancelled_magnetic_data: NDArray[(2, Any), int] = noise_cancellation(
            self.sensitivity_calibrated_tissue_gmr,
            self.sensitivity_calibrated_reference_gmr,
        )

        self.fully_calibrated_magnetic_data: NDArray[(2, Any), int] = apply_empty_plate_calibration(
            self.noise_cancelled_magnetic_data
        )

        if self.noise_filter_uuid is None:
            self.noise_filtered_magnetic_data: NDArray[(2, Any), int] = self.fully_calibrated_magnetic_data
        else:
            self.noise_filtered_magnetic_data: NDArray[(2, Any), int] = apply_noise_filtering(
                self.fully_calibrated_magnetic_data,
                self.filter_coefficients,
            )

        self.compressed_magnetic_data: NDArray[(2, Any), int] = compress_filtered_magnetic_data(
            self.noise_filtered_magnetic_data
        )
        self.compressed_voltage: NDArray[(2, Any), np.float32] = calculate_voltage_from_gmr(
            self.compressed_magnetic_data
        )
        self.compressed_displacement: NDArray[(2, Any), np.float32] = calculate_displacement_from_voltage(
            self.compressed_voltage
        )
        self.compressed_force: NDArray[(2, Any), np.float32] = calculate_force_from_displacement(
            self.compressed_displacement
        )

        self.voltage: NDArray[(2, Any), np.float32] = calculate_voltage_from_gmr(
            self.noise_filtered_magnetic_data
        )
        self.displacement: NDArray[(2, Any), np.float32] = calculate_displacement_from_voltage(self.voltage)
        self.force: NDArray[(2, Any), np.float32] = calculate_force_from_displacement(self.displacement)

    def get(self, key, default):
        try:
            return self[key]
        except:
            return default

    def __contains__(self, key):
        key = str(key) if isinstance(key, uuid.UUID) else key
        return key in self.attrs

    def __setitem__(self, key, newvalue):
        key = str(key) if isinstance(key, uuid.UUID) else key
        self.attrs[key] = newvalue

    def __getitem__(self, i):
        i = str(i) if isinstance(i, uuid.UUID) else i
        return self.attrs[i]

    def _extract_datetime(self, metadata_uuid: uuid.UUID) -> datetime.datetime:
        if self.version.split(".") < VersionInfo.parse("0.2.1"):
            if metadata_uuid == UTC_BEGINNING_RECORDING_UUID:
                """The use of this proxy value is justified by the fact that
                there is a 15 second delay between when data is recorded and
                when the GUI displays it, and because the GUI will send the
                timestamp of when the recording button is pressed."""
                acquisition_timestamp_str = self[UTC_BEGINNING_DATA_ACQUISTION_UUID]

                begin_recording = datetime.datetime.strptime(
                    acquisition_timestamp_str, DATETIME_STR_FORMAT
                ).replace(tzinfo=datetime.timezone.utc) + datetime.timedelta(seconds=15)

                return begin_recording
            if metadata_uuid == UTC_FIRST_TISSUE_DATA_POINT_UUID:
                """Early file versions did not include this metadata under a
                UUID, so we have to use this string identifier instead."""
                metadata_name = "UTC Timestamp of Beginning of Recorded Tissue Sensor Data"
                timestamp_str = self[metadata_name]

                return datetime.datetime.strptime(timestamp_str, DATETIME_STR_FORMAT).replace(
                    tzinfo=datetime.timezone.utc
                )
            if metadata_uuid == UTC_FIRST_REF_DATA_POINT_UUID:
                """Early file versions did not include this metadata under a
                UUID, so we have to use this string identifier instead."""
                timestamp_str = self["UTC Timestamp of Beginning of Recorded Reference Sensor Data"]

                return datetime.datetime.strptime(timestamp_str, DATETIME_STR_FORMAT).replace(
                    tzinfo=datetime.timezone.utc
                )

        timestamp_str = self[metadata_uuid]
        return datetime.datetime.strptime(timestamp_str, DATETIME_STR_FORMAT).replace(
            tzinfo=datetime.timezone.utc
        )

    def _load_reading(self, reading_type: str, time_trimmed) -> NDArray[(Any, Any), int]:
        recording_start_index = self[START_RECORDING_TIME_INDEX_UUID]
        beginning_data_acquisition_ts = self[UTC_BEGINNING_DATA_ACQUISTION_UUID]

        if reading_type == REFERENCE_SENSOR_READINGS:
            initial_timestamp = self[UTC_FIRST_REF_DATA_POINT_UUID]
            sampling_period = self[REF_SAMPLING_PERIOD_UUID]
        else:
            initial_timestamp = self[UTC_FIRST_TISSUE_DATA_POINT_UUID]
            sampling_period = self[TISSUE_SAMPLING_PERIOD_UUID]

        recording_start_index_useconds = int(recording_start_index) * MICROSECONDS_PER_CENTIMILLISECOND
        timestamp_of_start_index = beginning_data_acquisition_ts + datetime.timedelta(
            microseconds=recording_start_index_useconds
        )

        time_delta = initial_timestamp - timestamp_of_start_index
        time_delta_centimilliseconds = int(
            time_delta / datetime.timedelta(microseconds=MICROSECONDS_PER_CENTIMILLISECOND)
        )

        time_step = int(sampling_period / MICROSECONDS_PER_CENTIMILLISECOND)

        # adding `[:]` loads the data as a numpy array giving us more flexibility of multi-dimensional arrays
        data = self.file[reading_type][:]
        if len(data.shape) == 1:
            data = data.reshape(1, data.shape[0])

        # fmt: off
        # black reformatted this into a very ugly few lines of code
        times = np.mgrid[: data.shape[1],] * time_step
        # fmt: on

        if time_trimmed:
            new_times = times + time_delta_centimilliseconds
            start_index = _find_start_index(time_trimmed, new_times)
            time_delta_centimilliseconds = int(new_times[start_index])

        return np.concatenate(  # pylint: disable=unexpected-keyword-arg # Tanner (5/6/21): unsure why pylint thinks dtype is an unexpected kwarg for np.concatenate
            (times + time_delta_centimilliseconds, data), dtype=np.int32
        )


class PlateRecording:
    def __init__(self, path):
        self.path = path
        self.wells = []
        self._iter = 0
        self.is_optical_recording = False

        if self.path.endswith(".zip"):
            zf = zipfile.ZipFile(self.path)
            files = [f for f in zf.namelist() if f.endswith(".h5")]
            self.wells = [None] * len(files)

            with tempfile.TemporaryDirectory() as tempdir:
                zf.extractall(path=tempdir, members=files)

                for f in files:
                    well_file = WellFile(os.path.join(tempdir, f))
                    self.wells[well_file[WELL_INDEX_UUID]] = well_file
        elif self.path.endswith(".xlsx"):  # optical file
            self.is_optical_recording = True
            well_file = WellFile(self.path)
            self.wells = [None] * (well_file[WELL_INDEX_UUID] + 1)
            self.wells[well_file[WELL_INDEX_UUID]] = well_file
        else:  # directory of .h5 files
            files = glob.glob(os.path.join(self.path, "*.h5"))
            self.wells = [None] * len(files)

            for hf in files:
                well_file = WellFile(hf)
                self.wells[well_file[WELL_INDEX_UUID]] = well_file

    @staticmethod
    def from_directory(path):
        basedir = os.path.dirname(path)

        # multi zip files
        for zf in glob.glob(os.path.join(basedir, "*.zip")):
            yield PlateRecording(zf)

        # multi optical files
        for of in glob.glob(os.path.join(basedir, "*.xlsx")):
            yield PlateRecording(of)

        # directory of .h5 files
        yield PlateRecording(basedir)

    def __iter__(self):
        self._iter = 0
        return self

    def __next__(self):
        while self._iter < len(self.wells):
            value = self.wells[self._iter]
            self._iter += 1
            if not value:
                continue
            return value
        else:
            raise StopIteration


def _find_start_index(from_start: int, old_data: NDArray[(1, Any), int]) -> int:
    start_index, time_from_start = 0, 0

    while start_index + 1 < len(old_data) and from_start >= time_from_start:
        time_from_start = old_data[start_index + 1] - old_data[0]
        start_index += 1

    return start_index - 1  # loop iterates 1 past the desired index, so subtract 1
