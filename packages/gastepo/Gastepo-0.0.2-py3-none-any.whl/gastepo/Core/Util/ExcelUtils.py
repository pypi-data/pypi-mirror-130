# -*- coding: utf-8 -*-

import os

import xlwt
from openpyxl import load_workbook
from openpyxl.styles import Font
from xlrd import open_workbook
from xlrd.biffh import XLRDError
from xlutils.copy import copy

from Gastepo.Core.Base.CustomException import DotTypeError
from Gastepo.Core.Base.CustomException import SheetTypeError
from Gastepo.Core.Util.LogUtils import logger


class ExcelReader(object):
    """
    读取excel文件中的内容。返回list。

    如：
    excel中内容为：
    | A  | B  | C  |
    | A1 | B1 | C1 |
    | A2 | B2 | C2 |

    如果 print(ExcelReader(excel, data_type=True).data)，输出结果：
    [{A: A1, B: B1, C:C1}, {A:A2, B:B2, C:C2}]

    如果 print(ExcelReader(excel, data_type=False).data)，输出结果：
    [[A,B,C], [A1,B1,C1], [A2,B2,C2]]

    可以指定sheet，通过index或者name：
    """

    def __init__(self,
                 excel_path,
                 sheet=0,
                 data_type=True):
        '''
        指定默认参数文件、表格、构造标识
        :param excel: Excel文件路径
        :param sheet: 默认表格0
        :param data_type: 默认True, 按字典方式构造
        '''
        excel_extension = str(os.path.splitext(excel_path)[1])
        try:
            if os.path.exists(excel_path):
                if excel_extension.lower() not in [".xls", ".xlsx"]:
                    raise DotTypeError
            else:
                raise FileNotFoundError
            self.data_type = data_type
            self._data = list()
            workbook = open_workbook(excel_path)
            logger.info('[Initial]：当前准备读取的EXCEL文件为：{}'.format(excel_path))
            if type(sheet) == int:
                self.r = workbook.sheet_by_index(sheet)
                logger.info('Sheet表格索引限定为：{}'.format(sheet))
            elif type(sheet) == str:
                self.r = workbook.sheet_by_name(sheet)
                logger.info('Sheet表格名称限定为：{}'.format(sheet))
            else:
                raise SheetTypeError
        except DotTypeError:
            logger.exception(
                '非法文件扩展名"{}", 当前仅支持xls、xlsx格式Excel文件.'.format(excel_extension))
        except FileNotFoundError:
            logger.exception('非法EXCEL文件路径！~ "{}"'.format(excel_path))
        except SheetTypeError:
            logger.exception('请输入<type int>或<type str>类型的sheet标识, 当前{0}为非法标识类型.'
                             .format(type(sheet)))
        except IndexError:
            logger.exception('当前指定的Sheet表格索引"{}"超出阈值.'.format(sheet))
        except XLRDError:
            logger.exception('当前指定的Sheet表格名称"{}"并不存在.'.format(sheet))
        except Exception:
            logger.exception("初始化Excel文件数据过程中发生异常，请检查！")

    @property
    def session(self):
        """
        获取Excel读对象。
        :return:
        """
        return self.r

    @property
    def generate_data(self):
        '''
        返回构造的sheet表单内数据
        :return:
        '''
        try:
            if self.data_type:
                title = self.r.row_values(0)
                for col in range(1, self.r.nrows):
                    self._data.append(dict(zip(title, self.r.row_values(col))))
            else:
                for col in range(0, self.r.nrows):
                    self._data.append(self.r.row_values(col))
            logger.info("[Success]：当前Excel文件已成功完成数据生成！")
            return self._data
        except Exception:
            logger.exception("初始化Excel文件数据过程中发生异常，请检查！")


class ExcelXlsWriter(object):
    """
    XLS格式EXCEL文件写入工具类。
    """
    # 字体颜色枚举
    __STYLE_DICT = {
        "orange": xlwt.easyxf(
            'font:height 200, color-index orange, bold off;align: wrap on, vert centre, horiz center'),
        "boldorange": xlwt.easyxf(
            'font:height 230, color-index orange, bold on;align: wrap on, vert centre, horiz center'),
        "red": xlwt.easyxf('font:height 230, color-index red, bold on;align: wrap on, vert centre, horiz center'),
        "green": xlwt.easyxf('font:height 230, color-index green, bold on;align: wrap on, vert centre, horiz center'),
        "gray": xlwt.easyxf('font:height 210, color-index gray80, bold on;align: wrap on, vert centre, horiz center'),
        "black": xlwt.easyxf('font:height 200, color-index black, bold off;align: wrap on, vert centre, horiz center')
    }

    def __init__(self,
                 excel_path,
                 sheet=0):
        '''
        指定默认参数文件、表格、构造标识
        :param excel: Excel文件路径(xls格式)
        :param sheet: 默认表格0
        '''
        excel_extension = str(os.path.splitext(excel_path)[1])
        try:
            if os.path.exists(excel_path):
                if excel_extension.lower() not in [".xls"]:
                    raise DotTypeError
            else:
                raise FileNotFoundError
            self.excel_path = excel_path
            init_workbook = open_workbook(excel_path, formatting_info=True)
            sheet_reader = init_workbook.sheet_by_index(sheet)
            self.row_count = sheet_reader.nrows
            self.column_count = sheet_reader.ncols
            self.workbook = copy(init_workbook)
            logger.info('[Initial]：当前准备同步的EXCEL文件为：{}'.format(excel_path))
            self.w = self.workbook.get_sheet(sheet)
            logger.info('当前Sheet表格限定为：{}'.format(sheet))
        except DotTypeError:
            logger.exception(
                '当前不支持文件扩展名"{}", 请指定xls格式的表格文件.'.format(excel_extension))
        except FileNotFoundError:
            logger.exception('非法EXCEL文件路径！~ "{}"'.format(excel_path))

    @property
    def session(self):
        """
        获取Excel写对象(xls格式)。
        :return:
        """
        return self.w

    def write(self, row, col, value, style="black"):
        """
        写入单元格并设置字体颜色。
        :param row: 行索引
        :param col: 列索引
        :param value: 赋值
        :param style: 字体颜色样式（默认黑色）
        :return:
        """
        if str(style).lower() not in ["orange", "green", "red", "boldorange", "gray", "black"]:
            logger.warning(
                '当前字体样式不支持"{}"颜色，请选择如下字体颜色["orange", "green", "red", "boldorange", "gray", "black"]'.format(style))
        self.w.write(row, col, value, self.__STYLE_DICT.get(style))
        return self

    def write_row(self, row, value, style="black"):
        """
        整行赋值并设置字体颜色。
        :param row: 行索引
        :param value: 赋值
        :param style: 字体颜色样式（默认黑色）
        :return:
        """
        if str(style).lower() not in ["orange", "green", "red", "boldorange", "gray", "black"]:
            logger.warning(
                '当前字体样式不支持"{}"颜色，请选择如下字体颜色["orange", "green", "red", "boldorange", "gray", "black"]'.format(style))
        for col in range(self.column_count):
            self.write(row, col, value, style)
        return self

    def write_column(self, col, value, style="black"):
        """
        整列赋值并设置字体颜色。
        :param col: 列索引
        :param value: 赋值
        :param style: 字体颜色样式（默认黑色）
        :return:
        """
        if str(style).lower() not in ["orange", "green", "red", "boldorange", "gray", "black"]:
            logger.warning(
                '当前字体样式不支持"{}"颜色，请选择如下字体颜色["orange", "green", "red", "boldorange","gray", "black"]'.format(style))
        for row in range(1, self.row_count):
            self.write(row, col, value, style)
        return self

    def save(self):
        """
        写入后保存Excel文件(xls格式)。
        :return:
        """
        self.workbook.save(self.excel_path)


class ExcelXlsxWriter(object):
    """
    XLSX格式EXCEL文件写入工具类。
    """
    # 字体颜色枚举
    __STYLE_DICT = {
        "orange": Font(u'宋体', size=11, bold=False, italic=False, strike=False, color='FF7F00'),
        "boldorange": Font(u'宋体', size=12.5, bold=True, italic=False, strike=False, color='FF7F00'),
        "red": Font(u'宋体', size=12.5, bold=True, italic=False, strike=False, color='FF0000'),
        "green": Font(u'宋体', size=12.5, bold=True, italic=False, strike=False, color='228B22'),
        "gray": Font(u'宋体', size=11.5, bold=True, italic=False, strike=False, color='696969'),
        "black": Font(u'宋体', size=11, bold=False, italic=False, strike=False, color='000000'),
    }

    def __init__(self, excel_path, sheet=0):
        '''
        指定默认参数文件、表格、构造标识
        :param excel: Excel文件路径(xlsx格式)
        :param sheet: 默认表格0
        '''
        excel_extension = str(os.path.splitext(excel_path)[1])
        try:
            if os.path.exists(excel_path):
                if excel_extension.lower() not in [".xlsx"]:
                    raise DotTypeError
            else:
                raise FileNotFoundError
            self.excel_path = excel_path
            self.init_workbook = load_workbook(excel_path)
            specify_sheet = self.init_workbook.sheetnames[sheet]
            self.sheet_reader = self.init_workbook[specify_sheet]
            self.row_count = self.sheet_reader.max_row
            self.column_count = self.sheet_reader.max_column
            logger.info('[Initial]：当前准备同步的EXCEL文件为：{}'.format(excel_path))
            logger.info('当前Sheet表格限定为：{}'.format(sheet))
        except DotTypeError:
            logger.exception(
                '当前不支持文件扩展名"{}", 请指定xlsx格式的表格文件.'.format(excel_extension))
        except FileNotFoundError:
            logger.exception('非法EXCEL文件路径！~ "{}"'.format(excel_path))

    @property
    def session(self):
        """
        获取Excel写对象(xlsx格式)。
        :return:
        """
        return self.sheet_reader

    def write(self, row, col, value, style="black"):
        """
        写入单元格并设置字体颜色。
        :param row: 行索引
        :param col: 列索引
        :param value: 赋值
        :param style: 字体颜色样式（默认黑色）
        :return:
        """
        if str(style).lower() not in ["orange", "green", "red", "boldorange", "gray", "black"]:
            logger.warning(
                '当前字体样式不支持"{}"颜色，请选择如下字体颜色["orange", "green", "red", "boldorange", "gray", "black"]'.format(style))
        self.sheet_reader.cell(row + 1, col + 1).value = value
        self.sheet_reader.cell(row + 1, col + 1).font = self.__STYLE_DICT[style]
        return self

    def write_row(self, row, value, style="black"):
        """
        整行赋值并设置字体颜色。
        :param row: 行索引
        :param value: 赋值
        :param style: 字体颜色样式（默认黑色）
        :return:
        """
        if str(style).lower() not in ["orange", "green", "red", "boldorange", "gray", "black"]:
            logger.warning(
                '当前字体样式不支持"{}"颜色，请选择如下字体颜色["orange", "green", "red", "boldorange", "gray", "black"]'.format(style))
        for col in range(1, self.column_count + 1):
            self.sheet_reader.cell(row + 1, col).value = value
            self.sheet_reader.cell(row + 1, col).font = self.__STYLE_DICT[style]
        return self

    def write_column(self, col, value, style="black"):
        """
        整列赋值并设置字体颜色。
        :param col: 列索引
        :param value: 赋值
        :param style: 字体颜色样式（默认黑色）
        :return:
        """
        if str(style).lower() not in ["orange", "green", "red", "boldorange", "gray", "black"]:
            logger.warning(
                '当前字体样式不支持"{}"颜色，请选择如下字体颜色["orange", "green", "red", "boldorange","gray", "black"]'.format(style))
        for row in range(2, self.row_count + 1):
            self.sheet_reader.cell(row, col + 1).value = value
            self.sheet_reader.cell(row, col + 1).font = self.__STYLE_DICT[style]
        return self

    def save(self):
        """
        写入后保存Excel文件(xlsx格式)。
        :return:
        """
        self.init_workbook.save(self.excel_path)


if __name__ == '__main__':
    writer = ExcelXlsWriter(
        excel_path="/Users/mayer/PycharmProjects/AutoTest/Common/Data/Mente/ApiTestCase_stg.xls", sheet=0)
    # writer.write(1, 11, "Pass", style="green").save()
    # writer.write_column(15, "a", "red").save()
    # writer.write_row(3,"e","orange").save()
