# coding=utf-8

import pandas as pd
import mammut
import logging
from mammut.common.storage.storage_base import SheetsStorageBase
from openpyxl import load_workbook

logger = logging.getLogger()


class EmbeddedStorageExcel(SheetsStorageBase):

    DEFAULT_TOKEN_COLUMN_SEPARATOR = ":"
    DEFAULT_ENGINE = "openpyxl"
    DEFAULT_MAX_ROW = 940

    def __init__(self, storage_path):
        SheetsStorageBase.__init__(self, storage_path)

    def get_spreadsheet_as_dataframe(
        self,
        spreadsheet_id: str,
        sheet_name,
        start_column: str,
        start_row: str,
        end_range: str,
        transpose: bool = False,
    ):
        dataframe = None
        try:
            dataframe = pd.read_excel(
                self.storage_path + spreadsheet_id,
                sheet_name,
                usecols=start_column + self.DEFAULT_TOKEN_COLUMN_SEPARATOR + end_range,
                skiprows=int(start_row) - 1,
                keep_default_na=False,
                dtype=str,
            )
            if transpose:
                dataframe = dataframe.transpose()
        except Exception as ex:
            dataframe = None
            logger.error("Error getting sheet: " + str(ex))
            raise ex

        # This allows to fill sheets in a horizontal way like sheet 'non-functionals' from 'Standard4'
        dataframe.rename(columns=lambda x: x.split(".")[0], inplace=True)
        dataframe = dataframe.groupby(dataframe.columns, axis=1).first()

        return dataframe

    def add_rows_to_spreadsheet(
        self,
        spreadsheet_id: str,
        sheet_name: str,
        rows,
        start_row=SheetsStorageBase.DEFAULT_START_ROW,
        start_col=SheetsStorageBase.DEFAULT_START_COLUMN,
        end_range=SheetsStorageBase.DEFAULT_END_RANGE,
    ):

        try:
            book = load_workbook(self.storage_path + spreadsheet_id)
            ws = book.get_sheet_by_name(sheet_name)

            for i in range(int(start_row), self.DEFAULT_MAX_ROW):
                f = True
                for j in range(0, len(rows[0])):
                    val = ws.cell(row=i, column=j + 1).value
                    if val is not None:
                        f = False
                        break
                if f:
                    for h in range(0, len(rows)):
                        for k in range(0, len(rows[h])):
                            ws.cell(row=i + h, column=k + 1).value = rows[h][k]
                    break

            book.save(self.storage_path + spreadsheet_id)

        except Exception as ex:
            logger.error("Error adding rows on sheet: " + str(ex))
            return None

    def append_row_to_sheet(
        self,
        spreadsheet_id: str,
        sheet_name: str,
        row,
        start_row=SheetsStorageBase.DEFAULT_START_ROW,
        start_col=SheetsStorageBase.DEFAULT_START_COLUMN,
        end_range=SheetsStorageBase.DEFAULT_END_RANGE,
    ):
        try:
            book = load_workbook(self.storage_path + spreadsheet_id)
            ws = book.get_sheet_by_name(sheet_name)

            for i in range(int(start_row), self.DEFAULT_MAX_ROW):
                f = True
                for j in range(0, len(row)):
                    val = ws.cell(row=i, column=j + 1).value
                    if val is not None:
                        f = False
                        break
                if f:
                    for k in range(0, len(row)):
                        ws.cell(row=i, column=k + 1).value = row[k]
                    break

            book.save(self.storage_path + spreadsheet_id)
            return None
        except Exception as ex:
            logger.error("Error adding rows on sheet: " + str(ex))
            return None

    def create_new_sheet(self, spreadsheet_id: str, sheet_name: str):
        book = load_workbook(self.storage_path + spreadsheet_id)
        writer = pd.ExcelWriter(
            self.storage_path + spreadsheet_id, engine=self.DEFAULT_ENGINE
        )
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        pd.DataFrame().to_excel(writer, sheet_name)
        writer.save()

    def create_temporal_new_sheet(self, spreadsheet_id: str, base_sheet_title: str):
        sheet_title = (
            base_sheet_title + "-" + mammut.get_str_format_timezone_aware_datetime_now()
        )
        sheet_title = sheet_title.replace(
            mammut.TIME_SEPARATOR, mammut.TIME_ALTERNATIVE_SEPARATOR
        )
        self.create_new_sheet(spreadsheet_id, sheet_title)
        return sheet_title

    def get_temporal_sheets_from_new_to_old(self, spreadsheet_id, base_sheet_title):
        xl = pd.ExcelFile(self.storage_path + spreadsheet_id)
        titles = xl.sheet_names
        result = []
        for title in titles:
            if base_sheet_title in title and len(base_sheet_title) != len(title):
                time_stamp_str = title[len(base_sheet_title) + 1 :]
                time_stamp_str = time_stamp_str.replace(
                    mammut.TIME_ALTERNATIVE_SEPARATOR, mammut.TIME_SEPARATOR
                )
                time_stamp = mammut.str_parse_datetime(time_stamp_str)
                result.append((hash(title + spreadsheet_id), title, time_stamp))
        result = sorted(result, key=lambda i: i[2], reverse=True)
        return result

    def get_sheet_id(self, spreadsheet_id, identifier):
        return hash(identifier + spreadsheet_id)

    def update_row_from_spreadsheet(
        self, spreadsheet_id, sheet_name, row, start_row, start_col
    ):
        book = load_workbook(self.storage_path + spreadsheet_id)
        ws = book.get_sheet_by_name(sheet_name)
        for i in range(0, len(row)):
            ws.cell(row=int(start_row), column=i + 1).value = row[i]
        book.save(self.storage_path + spreadsheet_id)

    def update_rows_from_spreadsheet(
        self, spreadsheet_id, sheet_name, rows, start_row, start_col
    ):
        book = load_workbook(self.storage_path + spreadsheet_id)
        ws = book.get_sheet_by_name(sheet_name)
        for r, row in enumerate(rows):
            for i in range(0, len(row)):
                ws.cell(row=int(start_row) + r, column=i + 1).value = row[i]
        book.save(self.storage_path + spreadsheet_id)

    def delete_rows_from_spreadsheet(
        self, spreadsheet_id, sheet_name, start_index, end_index
    ):
        book = load_workbook(self.storage_path + spreadsheet_id)
        ws = book.get_sheet_by_name(sheet_name)
        ws.delete_rows(int(start_index), int(end_index) - int(start_index) + 1)
        book.save(self.storage_path + spreadsheet_id)

    def delete_sheet(self, spreadsheet_id, sheet_name):
        book = load_workbook(self.storage_path + spreadsheet_id)
        std = book.get_sheet_by_name(sheet_name)
        book.remove_sheet(std)
        book.save(self.storage_path + spreadsheet_id)
